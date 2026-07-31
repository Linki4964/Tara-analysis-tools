"""Fill the TARA template .xlsx with analysis results.

Template path: ``temples/模板.xlsx`` (relative to repo root).
"""

from __future__ import annotations

import io
import os
from typing import Any

import openpyxl

# Path to the template, resolved from this file's location
_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "..", "temples", "模板.xlsx")


def export_to_excel(
    project_name: str,
    assets: list[dict[str, Any]],
    threats: list[dict[str, Any]],
    attack_paths: list[dict[str, Any]],
    risk_treatments: list[dict[str, Any]],
    item_abbreviation: str = "VIU",
) -> bytes:
    """Fill the template with TARA results and return the workbook bytes."""
    wb = openpyxl.load_workbook(_TEMPLATE_PATH)

    _fill_cover(wb, project_name, item_abbreviation)
    _fill_asset_description(wb, assets, item_abbreviation)
    _fill_impact_analysis(wb, assets)
    _fill_threat_analysis(wb, threats, assets)
    _fill_attack_path_analysis(wb, attack_paths, threats, assets)
    _fill_risk_treatment(wb, risk_treatments, attack_paths)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _prepare_area(ws, start_row: int, max_row: int, min_col: int, max_col: int) -> None:
    """Unmerge and clear a rectangular area so we can write into it safely."""
    _unmerge_area(ws, start_row, max_row, min_col, max_col)
    for row in range(start_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).value = None


def _val(d: dict[str, Any], primary: str, fallback: str) -> Any:
    """Get ``d[primary]`` if present, else ``d[fallback]``, handling 0 correctly."""
    if primary in d and d[primary] is not None:
        return d[primary]
    return d.get(fallback)


def _unmerge_area(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    """Remove any merged cell ranges that overlap the given rectangular area."""
    to_remove = []
    for mc in ws.merged_cells.ranges:
        if mc.max_row < min_row or mc.min_row > max_row:
            continue
        if mc.max_col < min_col or mc.min_col > max_col:
            continue
        to_remove.append(str(mc))
    for mc_str in to_remove:
        ws.unmerge_cells(mc_str)


# ---------------------------------------------------------------------------
# Sheet fillers
# ---------------------------------------------------------------------------

def _fill_cover(wb, _project_name: str, item_abbr: str) -> None:
    ws = wb["1、封面"]
    ws["A9"] = f"for <{item_abbr}>"


def _fill_asset_description(wb, assets: list[dict[str, Any]], item_abbr: str) -> None:
    """Sheet 2 — columns B-G starting row 9."""
    ws = wb["2、网络安全相关项描述"]
    start_row = 9  # first data row
    max_row = ws.max_row

    _prepare_area(ws, start_row, max_row, 2, 6)

    for i, asset in enumerate(assets):
        row = start_row + i
        ws.cell(row=row, column=2, value=asset.get("assetName", ""))
        ws.cell(row=row, column=3, value=asset.get("assetId", ""))
        ws.cell(row=row, column=4, value=asset.get("assetType", ""))
        ws.cell(row=row, column=5, value=item_abbr)
        ws.cell(row=row, column=6, value=asset.get("description", ""))


def _fill_impact_analysis(wb, assets: list[dict[str, Any]]) -> None:
    """Sheet 3 — 影响分析, columns A-O starting row 4.  Each damage scenario = one row."""
    ws = wb["3、影响分析"]
    start_row = 4
    max_row = ws.max_row

    _prepare_area(ws, start_row, max_row, 1, 15)

    row = start_row
    for asset in assets:
        asset_ref = f"{asset.get('assetId', '')} {asset.get('assetName', '')}"
        asset_type = asset.get("assetType", "")

        for scenario in asset.get("damageScenarios") or []:
            ws.cell(row=row, column=1, value=asset_ref)
            ws.cell(row=row, column=2, value=asset_type)
            ws.cell(row=row, column=3, value=scenario.get("affectedProperty", ""))
            ws.cell(row=row, column=4, value=scenario.get("scenarioId", ""))
            ws.cell(row=row, column=5, value=scenario.get("description", ""))
            # S/F/O/P scores + rationales (use _val to handle 0 correctly)
            ws.cell(row=row, column=6, value=_val(scenario, "safetyScore", "safety"))
            ws.cell(row=row, column=7, value=scenario.get("safetyRationale", ""))
            ws.cell(row=row, column=8, value=_val(scenario, "financialScore", "financial"))
            ws.cell(row=row, column=9, value=scenario.get("financialRationale", ""))
            ws.cell(row=row, column=10, value=_val(scenario, "operationalScore", "operational"))
            ws.cell(row=row, column=11, value=scenario.get("operationalRationale", ""))
            ws.cell(row=row, column=12, value=_val(scenario, "privacyScore", "privacy"))
            ws.cell(row=row, column=13, value=scenario.get("privacyRationale", ""))
            ws.cell(row=row, column=14, value=scenario.get("damageImpactTotal"))
            ws.cell(row=row, column=15, value=scenario.get("damageImpactLevelLabel") or scenario.get("damageImpactLevel", ""))
            row += 1


def _fill_threat_analysis(wb, threats: list[dict[str, Any]], assets: list[dict[str, Any]]) -> None:
    """Sheet 4 — 威胁分析, columns A-D starting row 3."""
    ws = wb["4、威胁分析"]
    start_row = 3
    max_row = ws.max_row

    _prepare_area(ws, start_row, max_row, 1, 4)

    # asset-id → name lookup
    name_map = {a.get("assetId", ""): a.get("assetName", "") for a in assets}

    for i, threat in enumerate(threats):
        row = start_row + i
        target_id = threat.get("targetAsset", "")
        target_name = threat.get("targetAssetName") or name_map.get(target_id, "")
        asset_ref = f"{target_id} {target_name}".strip()

        ws.cell(row=row, column=1, value=asset_ref)
        ws.cell(row=row, column=2, value=threat.get("strideCategory", ""))
        ws.cell(row=row, column=3, value=threat.get("threatId", ""))
        ws.cell(row=row, column=4, value=threat.get("description", ""))


def _fill_attack_path_analysis(
    wb,
    attack_paths: list[dict[str, Any]],
    threats: list[dict[str, Any]],
    assets: list[dict[str, Any]],
) -> None:
    """Sheet 5 — 攻击路径分析, columns A-Q starting row 4."""
    ws = wb["5、攻击路径分析"]
    start_row = 4
    max_row = ws.max_row

    _prepare_area(ws, start_row, max_row, 1, 17)

    threat_map = {t.get("threatId", ""): t for t in threats}
    name_map = {a.get("assetId", ""): a.get("assetName", "") for a in assets}

    for i, ap in enumerate(attack_paths):
        row = start_row + i

        # Asset reference from first related threat
        related = ap.get("relatedThreats") or []
        first_t = threat_map.get(related[0]) if related else {}
        target_id = first_t.get("targetAsset", "")
        target_name = first_t.get("targetAssetName") or name_map.get(target_id, "")
        asset_ref = f"{target_id} {target_name}".strip()

        # Build description from entry + steps + consequence
        desc_parts = []
        if ap.get("entryPoint"):
            desc_parts.append(f"[入口] {ap['entryPoint']}")
        for step in (ap.get("attackSteps") or []):
            desc_parts.append(f"→ {step}")
        if ap.get("consequence"):
            desc_parts.append(f"[后果] {ap['consequence']}")

        threat_refs = ", ".join(related)

        # 5-dimension scores (prefer the structured dimensions dict, fall back to flat keys)
        dims = ap.get("attackFeasibilityDimensions") or {}
        et = dims.get("ET", {})
        exp = dims.get("EXP", {})
        kn = dims.get("KN", {})
        wo = dims.get("WO", {})
        eq = dims.get("EQ", {})

        et_score = et.get("score") if et else ap.get("et")
        et_rat = et.get("rationale", "") if et else ap.get("etRationale", "")
        exp_score = exp.get("score") if exp else ap.get("exp")
        exp_rat = exp.get("rationale", "") if exp else ap.get("expRationale", "")
        kn_score = kn.get("score") if kn else ap.get("kn")
        kn_rat = kn.get("rationale", "") if kn else ap.get("knRationale", "")
        wo_score = wo.get("score") if wo else ap.get("wo")
        wo_rat = wo.get("rationale", "") if wo else ap.get("woRationale", "")
        eq_score = eq.get("score") if eq else ap.get("eq")
        eq_rat = eq.get("rationale", "") if eq else ap.get("eqRationale", "")
        total = ap.get("attackFeasibilityTotal", (et_score or 0)+(exp_score or 0)+(kn_score or 0)+(wo_score or 0)+(eq_score or 0))
        al = ap.get("attackFeasibilityLabel") or ap.get("attackFeasibility", "")

        ws.cell(row=row, column=1, value=asset_ref)
        ws.cell(row=row, column=2, value=ap.get("attackPathId", ""))
        ws.cell(row=row, column=3, value=ap.get("relatedDamageScenarioId", ""))
        ws.cell(row=row, column=4, value=threat_refs)
        ws.cell(row=row, column=5, value="\n".join(desc_parts))
        ws.cell(row=row, column=6, value=et_score)
        ws.cell(row=row, column=7, value=et_rat)
        ws.cell(row=row, column=8, value=exp_score)
        ws.cell(row=row, column=9, value=exp_rat)
        ws.cell(row=row, column=10, value=kn_score)
        ws.cell(row=row, column=11, value=kn_rat)
        ws.cell(row=row, column=12, value=wo_score)
        ws.cell(row=row, column=13, value=wo_rat)
        ws.cell(row=row, column=14, value=eq_score)
        ws.cell(row=row, column=15, value=eq_rat)
        ws.cell(row=row, column=16, value=total)
        ws.cell(row=row, column=17, value=al)


def _fill_risk_treatment(
    wb,
    risk_treatments: list[dict[str, Any]],
    attack_paths: list[dict[str, Any]],
) -> None:
    """Sheet 6 — 风险处置与安全声明, columns A-K starting row 3."""
    ws = wb["6、风险处置与安全声明"]
    start_row = 3
    max_row = ws.max_row

    _prepare_area(ws, start_row, max_row, 1, 11)

    ap_map = {ap.get("attackPathId", ""): ap for ap in attack_paths}

    for i, rt in enumerate(risk_treatments):
        row = start_row + i
        ap = ap_map.get(rt.get("relatedAttackPath", ""), {})

        il = ap.get("impactLevelLabel") or ap.get("impactLevel", "")
        al = ap.get("attackFeasibilityLabel") or ap.get("attackFeasibility", "")
        sl_val = rt.get("securityRiskLevel", "")
        sl_label = rt.get("securityRiskLevelLabel", "")
        risk_rating = f"SL{sl_val} {sl_label}" if sl_val else ""
        decision = rt.get("treatmentDecisionLabel") or rt.get("treatmentDecision", "")

        ws.cell(row=row, column=1, value=rt.get("relatedDamageScenarioId", ""))
        ws.cell(row=row, column=2, value=rt.get("relatedThreatId", ""))
        ws.cell(row=row, column=3, value=il)
        ws.cell(row=row, column=4, value=al)
        ws.cell(row=row, column=5, value=risk_rating)
        ws.cell(row=row, column=6, value=decision)
        ws.cell(row=row, column=7, value=rt.get("cybersecurityGoalId", ""))
        ws.cell(row=row, column=8, value=rt.get("cybersecurityGoal", ""))
        ws.cell(row=row, column=9, value=rt.get("cybersecurityRequirement", ""))
        ws.cell(row=row, column=10, value=rt.get("cybersecurityClaimId", ""))
        ws.cell(row=row, column=11, value=rt.get("cybersecurityClaim", ""))
